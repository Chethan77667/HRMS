from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from flask_socketio import SocketIO, emit
from io import BytesIO
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from utils.db import users, leaves, salaries, timetable, init_db
from bson.objectid import ObjectId
import os

app = Flask(__name__)
app.secret_key = os.urandom(24)
socketio = SocketIO(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

class User(UserMixin):
    def __init__(self, user_data):
        self.id = str(user_data['_id'])
        self.username = user_data['username']
        self.role = user_data['role']
        self.name = user_data.get('name', '')

@login_manager.user_loader
def load_user(user_id):
    user_data = users.find_one({"_id": ObjectId(user_id)})
    if user_data:
        return User(user_data)
    return None

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('lecturer_dashboard'))
    return render_template('landing.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user_data = users.find_one({"username": username})
        
        if user_data and bcrypt.check_password_hash(user_data['password'], password):
            user_obj = User(user_data)
            login_user(user_obj)
            if user_obj.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('lecturer_dashboard'))
        else:
            flash('Invalid username or password', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

from utils.auth import admin_required, lecturer_required
from datetime import datetime
from utils.timetable_processor import pdf_to_faculty_images, extract_timetable_structure
from difflib import get_close_matches
import json
import re
import csv

# Admin Routes
@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    stats = {
        "staff_count": users.count_documents({"role": "lecturer"}),
        "pending_leaves": leaves.count_documents({"status": "Pending"}),
        "timetable_entries": timetable.count_documents({})
    }
    # Only show recent PENDING leaves in the dashboard widget
    recent_leaves = list(leaves.find({"status": "Pending"}).sort("_id", -1).limit(5))

    # Pre-serialize recent leaves for use in inline JS (ObjectId is not JSON serializable)
    recent_leaves_serialized = [
        {
            "id": str(doc.get("_id")),
            "lecturer_name": doc.get("lecturer_name", ""),
            "type": doc.get("type", ""),
            "from_date": doc.get("from_date", ""),
            "to_date": doc.get("to_date", ""),
            "status": doc.get("status", ""),
        }
        for doc in recent_leaves
    ]

    return render_template(
        'admin/dashboard.html',
        stats=stats,
        recent_leaves=recent_leaves,
        recent_leaves_serialized=recent_leaves_serialized,
    )


@app.route('/admin/api/recent-leaves')
@login_required
@admin_required
def admin_api_recent_leaves():
    """
    Small JSON API for polling recent pending leaves on the dashboard
    (used for near real-time updates without a full page refresh).
    """
    items = []
    for doc in leaves.find({"status": "Pending"}).sort("_id", -1).limit(5):
        items.append({
            "id": str(doc.get("_id")),
            "lecturer_name": doc.get("lecturer_name", ""),
            "type": doc.get("type", ""),
            "from_date": doc.get("from_date", ""),
            "to_date": doc.get("to_date", ""),
            "status": doc.get("status", ""),
        })
    return jsonify(items)

@app.route('/admin/staff')
@login_required
@admin_required
def manage_staff():
    # Always show lecturers sorted by Staff ID (BBHCF001, BBHCF002, ...)
    all_staff = list(users.find({"role": "lecturer"}).sort("staff_id", 1))
    return render_template('admin/manage_staff.html', staff=all_staff)

@app.route('/admin/staff/new', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_staff_new():
    error = None
    form = {
        "staff_id": "",
        "name": "",
        "designation": "",
        "department": "",
        "category": "Teaching Faculty",
        "email": "",
        "username": "",
    }

    if request.method == 'POST':
        staff_id = (request.form.get('staff_id') or '').strip()
        name = (request.form.get('name') or '').strip()
        designation = (request.form.get('designation') or '').strip()
        department = (request.form.get('department') or '').strip()
        category = (request.form.get('category') or '').strip()
        email = (request.form.get('email') or '').strip()
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''

        # Auto-set username = staff_id if not provided
        if not username and staff_id:
            username = staff_id.lower()

        # Default password = "123456" if not provided
        if not password:
            password = "123456"

        form.update(
            staff_id=staff_id,
            name=name,
            designation=designation,
            department=department,
            category=category,
            email=email,
            username=username,
        )

        if not staff_id or not name or not designation or not department or not category:
            error = "Please fill all required fields."
        elif users.find_one({"staff_id": staff_id}):
            error = "This Staff ID already exists."
        elif username and users.find_one({"username": username}):
            error = "This username already exists."

        if not error:
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
            users.insert_one({
                "role": "lecturer",
                "staff_id": staff_id,
                "name": name,
                "designation": designation,
                "department": department,
                "category": category,
                "email": email,
                "username": username,
                "password": password_hash,
                "display_password": password,  # Store for admin display
                "created_date": datetime.now(),  # Store creation date
                "assigned_subjects": "",  # Initialize assigned subjects
            })
            return redirect(url_for('manage_staff'))

    return render_template('admin/staff_form.html', mode="create", form=form, error=error)

@app.route('/admin/staff/<id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_staff_edit(id):
    error = None
    staff_doc = users.find_one({"_id": ObjectId(id), "role": "lecturer"})
    if not staff_doc:
        return redirect(url_for('manage_staff'))

    form = {
        "staff_id": staff_doc.get("staff_id", ""),
        "name": staff_doc.get("name", ""),
        "designation": staff_doc.get("designation", ""),
        "department": staff_doc.get("department", ""),
        "category": staff_doc.get("category", "Teaching Faculty"),
        "email": staff_doc.get("email", ""),
        "username": staff_doc.get("username", ""),
    }

    if request.method == 'POST':
        staff_id = (request.form.get('staff_id') or '').strip()
        name = (request.form.get('name') or '').strip()
        designation = (request.form.get('designation') or '').strip()
        department = (request.form.get('department') or '').strip()
        category = (request.form.get('category') or '').strip()
        email = (request.form.get('email') or '').strip()
        username = (request.form.get('username') or '').strip()
        new_password = request.form.get('password') or ''

        # Auto-set username = staff_id if not provided
        if not username and staff_id:
            username = staff_id.lower()

        form.update(
            staff_id=staff_id,
            name=name,
            designation=designation,
            department=department,
            category=category,
            email=email,
            username=username,
        )

        if not staff_id or not name or not designation or not department or not category:
            error = "Please fill all required fields."
        else:
            existing_staff_id = users.find_one({"staff_id": staff_id, "_id": {"$ne": staff_doc["_id"]}})
            if existing_staff_id:
                error = "This Staff ID already exists."
            elif not username:
                error = "Username is required for lecturer login."
            else:
                existing_username = users.find_one({"username": username, "_id": {"$ne": staff_doc["_id"]}})
                if existing_username:
                    error = "This username already exists."

        if not error:
            update = {
                "staff_id": staff_id,
                "name": name,
                "designation": designation,
                "department": department,
                "category": category,
                "email": email,
                "username": username,
            }
            if new_password.strip():
                update["password"] = bcrypt.generate_password_hash(new_password).decode('utf-8')
                update["display_password"] = new_password  # Store for admin display

            users.update_one({"_id": staff_doc["_id"]}, {"$set": update})
            return redirect(url_for('manage_staff'))

    return render_template('admin/staff_form.html', mode="edit", form=form, error=error, staff_id=str(staff_doc["_id"]))

@app.route('/admin/staff/<id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_staff_delete(id):
    users.delete_one({"_id": ObjectId(id), "role": "lecturer"})
    return redirect(url_for('manage_staff'))


@app.route('/admin/staff/delete-all', methods=['POST'])
@login_required
@admin_required
def admin_staff_delete_all():
    result = users.delete_many({"role": "lecturer"})
    count = result.deleted_count
    return redirect(url_for('manage_staff', delete_all_success=count))

@app.route('/admin/staff/<id>/change-password', methods=['POST'])
@login_required
@admin_required
def admin_staff_change_password(id):
    staff_doc = users.find_one({"_id": ObjectId(id), "role": "lecturer"})
    if not staff_doc:
        return redirect(url_for('manage_staff'))
    
    new_password = request.form.get('new_password', '').strip()
    if not new_password:
        return redirect(url_for('manage_staff', password_error="Password cannot be empty."))
    
    password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
    # Store display_password for admin view (not secure but for display purposes)
    users.update_one({"_id": ObjectId(id)}, {"$set": {
        "password": password_hash,
        "display_password": new_password  # Store plain text for display only
    }})
    return redirect(url_for('manage_staff', password_updated='1', updated_name=staff_doc.get('name', 'lecturer')))

@app.route('/admin/staff/export-excel')
@login_required
@admin_required
def admin_staff_export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        all_staff = list(users.find({"role": "lecturer"}).sort("staff_id", 1))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Lecturers"
        
        # Headers - only essential adding information
        headers = ["Lecturer ID", "Name", "Username", "Password"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data rows - only essential information
        for staff in all_staff:
            lecturer_id = staff.get('staff_id', '')
            name = staff.get('name', '')
            username = staff.get('username', staff.get('staff_id', '').lower())
            password = staff.get('display_password', '123456')
            
            ws.append([
                lecturer_id,
                name,
                username,
                password
            ])
        
        # Auto-adjust column widths
        column_widths = {
            'A': 15,  # Lecturer ID
            'B': 35,  # Name
            'C': 15,  # Username
            'D': 15   # Password
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Create in-memory file
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        filename = f"Lecturer_Management_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        flash('openpyxl library not installed. Please install it: pip install openpyxl', 'error')
        return redirect(url_for('manage_staff'))
    except Exception as e:
        flash(f'Error exporting Excel: {str(e)}', 'error')
        return redirect(url_for('manage_staff'))

@app.route('/admin/staff/bulk-upload', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_staff_bulk_upload():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            return render_template('admin/bulk_upload.html', error="No file selected.")
        
        file = request.files['excel_file']
        if file.filename == '':
            return render_template('admin/bulk_upload.html', error="No file selected.")
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return render_template('admin/bulk_upload.html', error="Please upload a valid Excel file (.xlsx or .xls).")
        
        try:
            import openpyxl
            from openpyxl import load_workbook
            
            workbook = load_workbook(file)
            sheet = workbook.active
            
            # Expected columns: Staff ID, Name, Designation, Department, Category, Email, Username, Password
            headers = [cell.value for cell in sheet[1]]
            
            # Find column indices (accept Staff ID, Name/Faculty Name/Faculty for display name)
            col_map = {}
            for idx, header in enumerate(headers, start=1):
                header_lower = str(header).lower().strip() if header else ""
                if 'staff' in header_lower and 'id' in header_lower:
                    col_map['staff_id'] = idx
                elif header_lower in ('faculty name', 'name') or header_lower == 'faculty':
                    col_map['name'] = idx  # Faculty Name / Name for full display (e.g. Mr.Umesh)
                elif 'name' in header_lower and 'user' not in header_lower:
                    col_map['name'] = idx  # e.g. Staff Name, but not Username
                elif 'designation' in header_lower:
                    col_map['designation'] = idx
                elif 'department' in header_lower:
                    col_map['department'] = idx
                elif 'category' in header_lower:
                    col_map['category'] = idx
                elif 'email' in header_lower:
                    col_map['email'] = idx
                elif 'username' in header_lower:
                    col_map['username'] = idx
                elif 'password' in header_lower:
                    col_map['password'] = idx
            
            if 'staff_id' not in col_map or 'name' not in col_map:
                return render_template('admin/bulk_upload.html', error="Excel file must have 'Staff ID' and 'Name' (or 'Faculty Name') columns.")
            
            success_count = 0
            error_rows = []
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    staff_id = str(row[col_map['staff_id'] - 1].value or '').strip()
                    name = str(row[col_map['name'] - 1].value or '').strip()
                    
                    if not staff_id or not name:
                        continue
                    
                    designation = str(row[col_map.get('designation', 0) - 1].value or '').strip() if col_map.get('designation') else ''
                    department = str(row[col_map.get('department', 0) - 1].value or '').strip() if col_map.get('department') else ''
                    category = str(row[col_map.get('category', 0) - 1].value or '').strip() if col_map.get('category') else 'Teaching Faculty'
                    email = str(row[col_map.get('email', 0) - 1].value or '').strip() if col_map.get('email') else ''
                    username = str(row[col_map.get('username', 0) - 1].value or '').strip() if col_map.get('username') else staff_id.lower()
                    password = str(row[col_map.get('password', 0) - 1].value or '').strip() if col_map.get('password') else '123456'
                    
                    # Check if staff_id already exists
                    if users.find_one({"staff_id": staff_id}):
                        error_rows.append(f"Row {row_num}: Staff ID {staff_id} already exists")
                        continue
                    
                    # Check if username already exists
                    if username and users.find_one({"username": username}):
                        error_rows.append(f"Row {row_num}: Username {username} already exists")
                        continue
                    
                    password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                    
                    users.insert_one({
                        "role": "lecturer",
                        "staff_id": staff_id,
                        "name": name,
                        "designation": designation,
                        "department": department,
                        "category": category,
                        "email": email,
                        "username": username,
                        "password": password_hash,
                        "display_password": password,  # Store for admin display
                        "created_date": datetime.now(),  # Store creation date
                        "assigned_subjects": "",  # Initialize assigned subjects
                    })
                    success_count += 1
                except Exception as e:
                    error_rows.append(f"Row {row_num}: {str(e)}")
            
            message = f"Successfully imported {success_count} record(s)."
            if error_rows:
                message += f" {len(error_rows)} error(s) occurred."
            
            # Redirect back to staff list with status so refresh is safe (no re-upload)
            return redirect(url_for('manage_staff', bulk_success=message))
        except ImportError:
            return render_template('admin/bulk_upload.html', error="openpyxl library not installed. Please install it: pip install openpyxl")
        except Exception as e:
            return render_template('admin/bulk_upload.html', error=f"Error processing file: {str(e)}")
    
    return render_template('admin/bulk_upload.html')

@app.route('/admin/leaves')
@login_required
@admin_required
def admin_leaves():
    # Optional filters: search query and month (YYYY-MM)
    q = (request.args.get("q") or "").strip()
    month = (request.args.get("month") or "").strip()

    all_leaves = list(leaves.find().sort("_id", -1))

    def matches_filters(doc):
        text_ok = True
        month_ok = True

        if q:
            q_lower = q.lower()
            text_fields = [
                str(doc.get("lecturer_name", "")),
                str(doc.get("type", "")),
                str(doc.get("reason", "")),
                str(doc.get("status", "")),
            ]
            text_ok = any(q_lower in field.lower() for field in text_fields)

        if month:
            # Expect month in format YYYY-MM, match against from_date / to_date strings
            from_date = str(doc.get("from_date", ""))
            to_date = str(doc.get("to_date", ""))
            month_ok = month in from_date or month in to_date

        return text_ok and month_ok

    filtered_leaves = [doc for doc in all_leaves if matches_filters(doc)]

    all_lecturers = list(users.find({"role": "lecturer"}).sort("name", 1))

    return render_template(
        'admin/leave_requests.html',
        leaves=filtered_leaves,
        q=q,
        month=month,
        lecturers=all_lecturers,
    )

@app.route('/admin/leaves/api/set_allocation/<id>', methods=['POST'])
@login_required
@admin_required
def api_set_leave_allocation(id):
    allocated = request.json.get('leaves_per_month', 1)
    try:
        allocated = float(allocated)
    except:
        allocated = 1
    users.update_one({"_id": ObjectId(id)}, {"$set": {"leaves_per_month": allocated}})
    return jsonify({"success": True})


@app.route('/admin/leaves/delete-all', methods=['POST'])
@login_required
@admin_required
def admin_leaves_delete_all():
    result = leaves.delete_many({})
    count = result.deleted_count
    flash(f"Deleted {count} leave record(s).", "success")
    return redirect(url_for('admin_leaves'))


@app.route('/admin/timetables', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_timetables():
    """
    Upload a PDF containing multiple individual faculty timetables.
    Each page is treated as one faculty timetable.
    """
    message = None
    error = None

    if request.method == 'POST':
        file = request.files.get('timetable_pdf')
        if not file or file.filename == '':
            error = "Please select a PDF file."
        elif not file.filename.lower().endswith('.pdf'):
            error = "Only PDF files are supported."
        else:
            try:
                pdf_bytes = file.read()

                # Build lecturer list and faculty details BEFORE OCR so we can pass
                # all known faculty names into the PDF processor (for pages that
                # don't explicitly contain a 'FACULTY:' label).
                all_lecturers = list(users.find({"role": "lecturer"}))

                faculty_details_path = os.path.join(os.path.dirname(__file__), "faculty_detail.json")
                faculty_details = {}
                if os.path.exists(faculty_details_path):
                    with open(faculty_details_path, encoding="utf-8") as f:
                        faculty_details = json.load(f)

                # Collect all known faculty display names (from JSON + DB)
                known_faculty_names = []
                for info in faculty_details.values():
                    name = (info.get("name") or "").strip()
                    if name:
                        known_faculty_names.append(name)
                for lect in all_lecturers:
                    name = (lect.get("name") or "").strip()
                    if name:
                        known_faculty_names.append(name)

                pages_with_name, pages_without_name = pdf_to_faculty_images(
                    pdf_bytes, known_faculty_names=known_faculty_names or None
                )

                def normalize_name(name: str) -> str:
                    """
                    Make names comparable between OCR, faculty_detail.json and DB.
                    - Uppercase
                    - Remove common titles (MR, MRS, MS, DR, PROF, PROF.)
                    - Remove dots and extra spaces
                    """
                    name = (name or "").upper()
                    # Remove "FACULTY:" prefix if it slipped through
                    name = re.sub(r"^FACULTY\s*[:\-]\s*", "", name)
                    # Remove titles
                    name = re.sub(r"\b(MR|MRS|MS|MISS|DR|PROF|PROFESSOR)\.?\b", "", name)
                    # Remove non-letters except spaces
                    name = re.sub(r"[^A-Z\s]", " ", name)
                    # Collapse spaces
                    name = re.sub(r"\s+", " ", name).strip()
                    return name

                def surname_key(norm_name: str) -> str:
                    """
                    Return the last word of a normalized name (surname/last token).
                    Used as a very forgiving fallback when full-name matching fails.
                    """
                    parts = (norm_name or "").split()
                    return parts[-1] if parts else ""

                def partial_match(norm_small: str, norm_big: str) -> bool:
                    """
                    Check if one normalized name is mostly contained in another.
                    Example: 'WILMA SHARAL' vs 'WILMA SHARAL CORNELIO'.
                    """
                    if not norm_small or not norm_big:
                        return False
                    # Ensure we compare the shorter against the longer
                    if len(norm_small) > len(norm_big):
                        norm_small, norm_big = norm_big, norm_small
                    # Direct substring check
                    if norm_small in norm_big:
                        return True
                    # Token-overlap check: at least 2 common tokens
                    small_tokens = set(norm_small.split())
                    big_tokens = set(norm_big.split())
                    return len(small_tokens & big_tokens) >= min(2, len(small_tokens))

                # Map staff_id -> lecturer document
                lecturers_by_staff_id = {
                    lect.get("staff_id"): lect for lect in all_lecturers if lect.get("staff_id")
                }

                # Map normalized lecturer name -> lecturer (direct DB lookup)
                lecturers_by_norm_name = {}
                lecturers_by_surname = {}
                for lect in all_lecturers:
                    norm = normalize_name(lect.get("name", ""))
                    if norm:
                        lecturers_by_norm_name[norm] = lect
                        sk = surname_key(norm)
                        if sk:
                            lecturers_by_surname.setdefault(sk, []).append(lect)

                # Map normalized name from JSON -> staff_id (only teaching faculty)
                norm_json_name_to_staff_id = {}
                for staff_id, info in faculty_details.items():
                    if info.get("category") != "Teaching Faculty":
                        continue
                    norm = normalize_name(info.get("name", ""))
                    if norm:
                        norm_json_name_to_staff_id[norm] = staff_id

                upload_folder = os.path.join('static', 'timetables')
                os.makedirs(upload_folder, exist_ok=True)

                matched_count = 0
                unmatched_pages = []

                for page in pages_with_name:
                    faculty_name_raw = page["faculty_name"]
                    if not faculty_name_raw:
                        continue

                    norm_ocr_name = normalize_name(faculty_name_raw)
                    lecturer = None
                    matched_display_name = None

                    # 1) Exact match with JSON names -> staff_id -> lecturer
                    staff_id = norm_json_name_to_staff_id.get(norm_ocr_name)
                    if staff_id:
                        lecturer = lecturers_by_staff_id.get(staff_id)

                    # 2) Exact match with lecturer.name after normalization
                    if not lecturer:
                        lecturer = lecturers_by_norm_name.get(norm_ocr_name)

                    # 3) Partial match: OCR name is subset/prefix of a longer stored name or vice versa
                    if not lecturer and lecturers_by_norm_name:
                        for norm_name, lect in lecturers_by_norm_name.items():
                            if partial_match(norm_ocr_name, norm_name):
                                lecturer = lect
                                break

                    # 4) Fuzzy match on normalized lecturer names (to allow small spelling mistakes)
                    if not lecturer and lecturers_by_norm_name:
                        norm_lecturer_names = list(lecturers_by_norm_name.keys())
                        best = get_close_matches(norm_ocr_name, norm_lecturer_names, n=1, cutoff=0.6)
                        if best:
                            lecturer = lecturers_by_norm_name.get(best[0])

                    # 5) Very loose surname-only fallback: if exactly one lecturer shares this surname
                    if not lecturer:
                        sk = surname_key(norm_ocr_name)
                        if sk:
                            candidates = lecturers_by_surname.get(sk, [])
                            if len(candidates) == 1:
                                lecturer = candidates[0]

                    if not lecturer:
                        unmatched_pages.append({
                            "faculty_name": faculty_name_raw,
                            "normalized_name": norm_ocr_name,
                            "page_index": page["page_index"],
                        })
                        continue

                    matched_display_name = lecturer.get("name", faculty_name_raw)

                    # Save image
                    safe_name = matched_display_name.replace(" ", "_").replace(".", "").lower()
                    filename = f"{safe_name}.png"
                    fs_image_path = os.path.join(upload_folder, filename)
                    page["image"].save(fs_image_path, format="PNG")

                    # Store a URL-friendly relative path (forward slashes) for static serving
                    url_image_path = f"timetables/{filename}"

                    # Optional: extract structured timetable slots via AI (best-effort)
                    structured = extract_timetable_structure(page["image"])

                    # Upsert timetable record
                    timetable.update_one(
                        {"lecturer_id": str(lecturer["_id"])},
                        {
                            "$set": {
                                "lecturer_id": str(lecturer["_id"]),
                                "lecturer_name": matched_display_name,
                                "image_path": url_image_path,
                                "structured": structured or {},
                                "uploaded_at": datetime.now(),
                            }
                        },
                        upsert=True,
                    )
                    matched_count += 1

                message = f"Processed {len(pages_with_name)} page(s). Matched {matched_count} timetable(s)."
                if pages_without_name or unmatched_pages:
                    extra = []
                    if pages_without_name:
                        extra.append(f"{len(pages_without_name)} page(s) without detectable faculty name")
                    if unmatched_pages:
                        extra.append(f"{len(unmatched_pages)} page(s) where faculty name did not match any lecturer")

                        # Show a small sample of what OCR actually read so admin can verify.
                        sample_names = []
                        seen = set()
                        for item in unmatched_pages:
                            key = item.get("faculty_name") or ""
                            if key and key not in seen:
                                seen.add(key)
                                norm_display = item.get("normalized_name") or ""
                                sample_names.append(f"'{key}' -> '{norm_display}'")
                            if len(sample_names) >= 5:
                                break
                        if sample_names:
                            extra.append("sample OCR names: " + "; ".join(sample_names))

                    message += " " + "; ".join(extra) + "."
            except Exception as e:
                error = f"Error processing PDF: {str(e)}"

    # For display: list all lecturers with timetable status
    all_lecturers = list(users.find({"role": "lecturer"}))
    timetable_docs = {doc.get("lecturer_id"): doc for doc in timetable.find({})}

    lecturer_rows = []
    uploaded_count = 0
    for lect in all_lecturers:
        lect_id = str(lect["_id"])
        tt_doc = timetable_docs.get(lect_id)
        has_tt = tt_doc is not None
        image_url = ""
        if has_tt:
            image_path = (tt_doc.get("image_path") or "").replace("\\", "/")
            image_url = url_for("static", filename=image_path)
        if has_tt:
            uploaded_count += 1
        lecturer_rows.append({
            "id": lect_id,
            "name": lect.get("name", ""),
            "staff_id": lect.get("staff_id", ""),
            "has_timetable": has_tt,
            "timetable": tt_doc,
            "timetable_image_url": image_url,
        })

    pending_count = len(all_lecturers) - uploaded_count

    return render_template(
        'admin/timetables.html',
        lecturers=lecturer_rows,
        uploaded_count=uploaded_count,
        pending_count=pending_count,
        total=len(all_lecturers),
        message=message,
        error=error,
    )

@app.route('/admin/leave/<id>/<status>', methods=['GET', 'POST'])
@login_required
@admin_required
def review_leave(id, status):
    if status not in ("Approved", "Rejected", "Pending"):
        flash("Invalid status.", "danger")
        return redirect(request.referrer or url_for('admin_leaves'))

    leave_doc = leaves.find_one({"_id": ObjectId(id)})

    leaves.update_one(
        {"_id": ObjectId(id)},
        {"$set": {"status": status, "reviewed_at": datetime.now()}},
    )
    
    if leave_doc:
        leaves_left = calculate_leaves_left(leave_doc['lecturer_id'])
        socketio.emit('leave_status_update', {
            'id': id,
            'status': status,
            'lecturer_id': leave_doc['lecturer_id'],
            'leaves_left': leaves_left
        })
        
    flash(f"Leave {status.lower()} successfully!", "success")
    return redirect(request.referrer or url_for('admin_leaves'))

@app.route('/admin/leave/api/<id>/<status>', methods=['POST'])
@login_required
@admin_required
def api_review_leave(id, status):
    if status not in ("Approved", "Rejected", "Pending"):
        return jsonify({"success": False, "message": "Invalid status"}), 400

    leave_doc = leaves.find_one({"_id": ObjectId(id)})
    if not leave_doc:
        return jsonify({"success": False, "message": "Not found"}), 404

    leaves.update_one(
        {"_id": ObjectId(id)},
        {"$set": {"status": status, "reviewed_at": datetime.now()}},
    )
    
    leaves_left = calculate_leaves_left(leave_doc['lecturer_id'])
    socketio.emit('leave_status_update', {
        'id': id,
        'status': status,
        'lecturer_id': leave_doc['lecturer_id'],
        'leaves_left': leaves_left
    })
    
    return jsonify({"success": True})

def calculate_leaves_left(lecturer_id):
    user_doc = users.find_one({"_id": ObjectId(lecturer_id)})
    total_leaves = user_doc.get("leaves_per_month", 20) if user_doc else 20
    approved_leaves = list(leaves.find({"lecturer_id": lecturer_id, "status": "Approved"}))
    used_days = 0
    for l in approved_leaves:
        mode = l.get('mode', 'full')
        if mode == 'time':
            used_days += 1
        else:
            try:
                f_date = datetime.strptime(l['from_date'].split(' ')[0], '%Y-%m-%d')
                t_date = datetime.strptime(l['to_date'].split(' ')[0], '%Y-%m-%d')
                days = (t_date - f_date).days + 1
                if days > 0:
                    used_days += days
            except Exception:
                used_days += 1
    return max(0, total_leaves - used_days)

# Lecturer Routes
@app.route('/lecturer/dashboard')
@login_required
@lecturer_required
def lecturer_dashboard():
    my_leaves = list(leaves.find({"lecturer_id": current_user.id}).sort("_id", -1).limit(5))

    tt_doc = timetable.find_one({"lecturer_id": current_user.id})
    timetable_image_url = None
    has_timetable = False
    if tt_doc and tt_doc.get("image_path"):
        image_path = (tt_doc.get("image_path") or "").replace("\\", "/")
        timetable_image_url = url_for("static", filename=image_path)
        has_timetable = True

    leaves_left = calculate_leaves_left(current_user.id)

    return render_template(
        'lecturer/dashboard.html',
        leaves=my_leaves,
        has_timetable=has_timetable,
        timetable_image_url=timetable_image_url,
        leaves_left=leaves_left
    )


@app.route('/lecturer/attendance')
@login_required
@lecturer_required
def lecturer_attendance():
    """
    Attendance view for the logged-in lecturer.
    Reads JSON attendance files from ATTENDANCE_DIR and filters by this lecturer's staff ID.
    """
    base_dir = (os.getenv("ATTENDANCE_DIR") or "").strip()
    from datetime import datetime

    # Filters
    selected_month = (request.args.get("month") or "").strip()
    search_q = (request.args.get("q") or "").strip().lower()

    if not selected_month:
        selected_month = datetime.now().strftime("%Y-%m")

    # Find staff_id for current lecturer
    staff_doc = users.find_one({"_id": ObjectId(current_user.id)})
    staff_id = staff_doc.get("staff_id") if staff_doc else None

    records = []
    debug_info = {
        "base_dir": base_dir,
        "dir_exists": os.path.isdir(base_dir) if base_dir else False,
        "staff_id": staff_id,
        "json_files": [],
        "total_rows_all_files": 0,
        "rows_for_staff_before_filters": 0,
    }
    today_str = datetime.now().strftime("%Y-%m-%d")

    if base_dir and debug_info["dir_exists"] and staff_id:
        for fname in os.listdir(base_dir):
            if not fname.lower().endswith(".json"):
                continue
            fpath = os.path.join(base_dir, fname)
            debug_info["json_files"].append(fname)

            try:
                with open(fpath, encoding="utf-8") as f:
                    # Try to load as a JSON array or object first
                    try:
                        data = json.load(f)
                        if isinstance(data, dict):
                            data = [data]
                    except json.JSONDecodeError:
                        # Fallback: newline-delimited JSON objects
                        f.seek(0)
                        data = []
                        for line in f:
                            line = line.strip()
                            if not line:
                                continue
                            try:
                                data.append(json.loads(line))
                            except Exception:
                                continue

                for row in data:
                    debug_info["total_rows_all_files"] += 1

                    if row.get("staff_id") != staff_id:
                        continue
                    debug_info["rows_for_staff_before_filters"] += 1

                    checkin = row.get("checkin") or ""
                    checkout = row.get("checkout") or ""
                    name = row.get("name") or ""

                    # Derive date and month from checkin
                    iso_date = ""
                    display_date = ""
                    time_in = ""
                    time_out = ""
                    if checkin:
                        try:
                            dt = datetime.fromisoformat(checkin)
                            iso_date = dt.date().isoformat()
                            display_date = dt.date().strftime("%d-%m-%Y")
                            time_in = dt.time().strftime("%H:%M")
                        except Exception:
                            # Fallback: first 10 chars as date, last 8 as time if possible
                            if len(checkin) >= 10:
                                iso_date = checkin[:10]
                                try:
                                    dparts = iso_date.split("-")
                                    if len(dparts) == 3:
                                        display_date = f"{dparts[2]}-{dparts[1]}-{dparts[0]}"
                                except Exception:
                                    display_date = iso_date
                            if len(checkin) >= 19:
                                time_in = checkin[11:16]

                    if checkout:
                        try:
                            dt_out = datetime.fromisoformat(checkout)
                            time_out = dt_out.time().strftime("%H:%M")
                        except Exception:
                            if len(checkout) >= 19:
                                time_out = checkout[11:16]

                    # Month filter based on iso_date (YYYY-MM)
                    if iso_date and not iso_date.startswith(selected_month):
                        continue

                    # Simple status from presence of checkin/checkout
                    if checkin and checkout:
                        status = "Present"
                    elif checkin:
                        status = "Checked-in"
                    else:
                        status = "Unknown"

                    extra = f"In: {checkin}  Out: {checkout}"
                    if name:
                        extra = f"{name} | " + extra

                    text_blob = f"{iso_date} {time_in} {time_out} {status} {extra}".lower()
                    if search_q and search_q not in text_blob:
                        continue

                    records.append({
                        "date": iso_date,
                        "display_date": display_date or iso_date,
                        "time_in": time_in,
                        "time_out": time_out,
                        "status": status,
                        "extra": extra,
                    })
            except Exception:
                continue

    # Sort by date+time descending
    def sort_key(rec):
        return (rec.get("date") or "", rec.get("time") or "")

    records.sort(key=sort_key, reverse=True)

    today_records = [r for r in records if r.get("date") == today_str]

    return render_template(
        "lecturer/attendance.html",
        records=records,
        today_records=today_records,
        month=selected_month,
        q=search_q,
        debug_info=debug_info,
    )

@app.route('/lecturer/apply-leave', methods=['GET', 'POST'])
@login_required
@lecturer_required
def apply_leave():
    mode = request.args.get('mode', 'full')
    if request.method == 'POST':
        leave_mode = request.form.get('mode', 'full')
        
        if leave_mode == 'time':
            today_str = datetime.now().strftime('%Y-%m-%d')
            time_from = request.form.get('time_from', '')
            time_to = request.form.get('time_to', '')
            from_date = f"{today_str} {time_from}"
            to_date = f"{today_str} {time_to}"
        else:
            from_date = request.form.get('from_date')
            to_date = request.form.get('to_date')

        leave_data = {
            "lecturer_id": current_user.id,
            "lecturer_name": current_user.name,
            "type": request.form.get('type'),
            "from_date": from_date,
            "to_date": to_date,
            "reason": request.form.get('reason'),
            "status": "Pending",
            "created_at": datetime.now(),
            "mode": leave_mode
        }
        res = leaves.insert_one(leave_data)
        
        socketio.emit('new_leave_request', {
            "id": str(res.inserted_id),
            "lecturer_name": current_user.name,
            "type": request.form.get('type'),
            "from_date": from_date,
            "to_date": to_date,
            "status": "Pending"
        })
        
        flash("Leave application submitted successfully!", "success")
        return redirect(url_for('lecturer_dashboard'))
    return render_template('lecturer/apply_leave.html', mode=mode)


@app.route('/lecturer/leave/<id>/cancel', methods=['POST'])
@login_required
@lecturer_required
def cancel_leave(id):
    """
    Allow a lecturer to cancel one of their own pending leave requests.
    """
    leave_doc = leaves.find_one({"_id": ObjectId(id), "lecturer_id": current_user.id})
    if not leave_doc:
        flash("Leave request not found.", "danger")
        return redirect(url_for('lecturer_dashboard'))

    if leave_doc.get("status") != "Pending":
        flash("Only pending leave requests can be cancelled.", "warning")
        return redirect(url_for('lecturer_dashboard'))

    leaves.update_one(
        {"_id": leave_doc["_id"]},
        {"$set": {"status": "Cancelled", "cancelled_at": datetime.now()}},
    )
    
    socketio.emit('leave_cancelled', {'id': id, 'lecturer_id': current_user.id})
    flash("Leave request cancelled.", "success")
    return redirect(url_for('lecturer_dashboard'))

@app.route('/lecturer/leave/api/<id>/cancel', methods=['POST'])
@login_required
@lecturer_required
def api_cancel_leave(id):
    leave_doc = leaves.find_one({"_id": ObjectId(id), "lecturer_id": current_user.id})
    if not leave_doc:
        return jsonify({"success": False, "message": "Not found"}), 404
        
    if leave_doc.get("status") != "Pending":
        return jsonify({"success": False, "message": "Not pending"}), 400
        
    leaves.update_one(
        {"_id": leave_doc["_id"]},
        {"$set": {"status": "Cancelled", "cancelled_at": datetime.now()}},
    )
    
    socketio.emit('leave_cancelled', {'id': id, 'lecturer_id': current_user.id})
    return jsonify({"success": True})

@app.route('/lecturer/salary')
@login_required
@lecturer_required
def view_salary():
    my_salaries = list(salaries.find({"lecturer_id": current_user.id}).sort("month_year", -1))
    return render_template('lecturer/salary.html', salaries=my_salaries)


@app.route('/lecturer/timetable')
@login_required
@lecturer_required
def lecturer_timetable():
    """Show the logged-in lecturer's own timetable image (uploaded by admin)."""
    tt_doc = timetable.find_one({"lecturer_id": current_user.id})
    image_url = None
    if tt_doc and tt_doc.get("image_path"):
        image_path = (tt_doc.get("image_path") or "").replace("\\", "/")
        image_url = url_for("static", filename=image_path)
    return render_template(
        'lecturer/timetable.html',
        has_timetable=image_url is not None,
        timetable_image_url=image_url,
        structured=tt_doc.get("structured") if tt_doc else {},
    )


@app.route('/lecturer/timetable/edit', methods=['GET', 'POST'])
@login_required
@lecturer_required
def edit_lecturer_timetable():
    """
    Simple JSON-editor for the structured timetable extracted by AI.
    This lets a lecturer tweak the parsed slots without changing the image.
    """
    tt_doc = timetable.find_one({"lecturer_id": current_user.id})
    if not tt_doc:
        flash("No timetable found to edit. Please contact administration.", "warning")
        return redirect(url_for('lecturer_timetable'))

    import json as _json

    structured = tt_doc.get("structured") or {}
    structured_text = _json.dumps(structured, indent=2, ensure_ascii=False)

    if request.method == 'POST':
        raw = request.form.get("structured_json", "").strip()
        if not raw:
            flash("Timetable JSON cannot be empty.", "danger")
            return redirect(url_for('edit_lecturer_timetable'))
        try:
            data = _json.loads(raw)
            if not isinstance(data, dict):
                raise ValueError("Root must be a JSON object.")
        except Exception as exc:
            flash(f"Invalid JSON: {exc}", "danger")
            return render_template(
                'lecturer/edit_timetable.html',
                structured_json=raw,
            )

        timetable.update_one(
            {"_id": tt_doc["_id"]},
            {"$set": {"structured": data}},
        )
        flash("Timetable updated.", "success")
        return redirect(url_for('lecturer_timetable'))

    return render_template(
        'lecturer/edit_timetable.html',
        structured_json=structured_text,
    )


if __name__ == '__main__':
    init_db()
    # On some Windows setups (especially with newer Python), the watchdog reloader can throw WinError 10038.
    # Disabling the reloader keeps dev runs stable; restart the server manually after code changes.
    socketio.run(app, debug=True, use_reloader=False)
