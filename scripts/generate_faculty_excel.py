"""
Generate Excel sheet from faculty_detail.json for HRMS bulk upload.
Run: python scripts/generate_faculty_excel.py
Output: faculty_upload.xlsx (in project root)
"""
import json
import os
import re
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)

# Paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent
JSON_PATH = PROJECT_ROOT / "faculty_detail.json"
OUTPUT_PATH = PROJECT_ROOT / "faculty_upload.xlsx"


def sanitize_for_filename(text):
    """Create a valid username/email part from name."""
    if not text:
        return ""
    # Remove titles and special chars
    text = re.sub(r"(Prof\.|Dr\.|Mr\.|Mrs\.|Ms\.)\s*", "", text, flags=re.I)
    text = re.sub(r"[^a-zA-Z0-9]", "", text)
    return text.lower()[:20] if text else ""


def main():
    if not JSON_PATH.exists():
        print(f"Error: {JSON_PATH} not found.")
        exit(1)

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty"

    # Headers matching bulk upload format
    headers = [
        "Staff ID",
        "Name",
        "Designation",
        "Department",
        "Category",
        "Email",
        "Username",
        "Password",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for staff_id, details in data.items():
        name = details.get("name", "")
        designation = details.get("designation", "")
        department = details.get("department", "")
        category = details.get("category", "Teaching Faculty")

        # Generate username: staff_id lowercase
        username = str(staff_id).lower()

        # Generate email: username@college.edu or derived from name
        base = sanitize_for_filename(name) or username
        email = f"{base}@college.edu"

        # Default password for login
        password = f"{staff_id.lower()}123"

        ws.append([
            staff_id,
            name,
            designation,
            department,
            category,
            email,
            username,
            password,
        ])

    wb.save(OUTPUT_PATH)
    print(f"Generated: {OUTPUT_PATH}")
    print(f"Total records: {len(data)}")


if __name__ == "__main__":
    main()
